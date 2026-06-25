"""Placeholder 3 (delivery report) engine.

This placeholder's own working engine. Shared low-level primitives
(Excel read/write, WCAG tag-file, audit header template) are imported
from services.feature_service so they are defined once.
"""
from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

import polars as pl

from config import config
from services import export_service as _ex
from services.excel_service import read_sheet

from services.feature_service import (
    TEMPLATE_DIR, WCAG_TEMPLATE, WCAG_DELIVERY_TEMPLATE, DIGITAL_ASSET_TEMPLATE,
    ALT_TEXT_TEMPLATE, WCAG_TAGS,
    list_sheets_for, template_status, audit_headers, validate_audit_columns,
    strip_description_hyphen, load_wcag_tags, space_description_sections,
    _space_one_description,
    EXCEL_ERROR_VALUES, _read_table, _find_col, _col_or_blank, _outpath, _ts,
    _slug, _preview, _exact_col, _category_from_issue, _excel_error_cells,
    _AUDIT_HEADERS_FALLBACK,
)
# Reuse Placeholder 4's document classifier (PDF / Word / PowerPoint from a
# Type/URL/Name blob) rather than re-implementing it here.
from dashboards.dashboard4_downloadable.engine import _classify_doc


def prepare_delivery_export(df: pl.DataFrame) -> tuple[pl.DataFrame, str | None]:
    """Apply Placeholder-3 export rules to a delivery frame (no ``__id`` column):

      3. Fill any blank 'Issue Category' from 'Issue' (hyphens -> spaces,
         Title Case).
      4. Remove a leading hyphen (and surrounding spaces) from every
         'Description' cell, so Excel doesn't treat the text as a formula.
      2. Reject the export when ANY cell is still blank, returning a message
         that names the offending columns/rows.

    Returns ``(transformed_df, error_message_or_None)``.
    """
    if df.height == 0:
        return df, "Nothing to export — the sheet has no rows."

    def _blank_expr(col: str) -> pl.Expr:
        return (pl.col(col).is_null()
                | (pl.col(col).cast(pl.Utf8, strict=False)
                     .str.strip_chars().fill_null("") == ""))

    # NOTE: the blank 'Issue Category' <- 'Issue' auto-fill was removed from
    # placeholder 3. That conversion condition now runs in placeholder 2
    # (axe -> audit), so Issue Category arrives already populated; any cell left
    # blank here is reported by the blank-cell check below like any other.

    # --- 2. block export if any cell is still blank ---------------------------
    blanks: dict[str, list[int]] = {}
    for c in df.columns:
        mask = df.select(_blank_expr(c).alias("b")).get_column("b").to_list()
        rows = [i + 1 for i, b in enumerate(mask) if b]
        if rows:
            blanks[c] = rows
    if blanks:
        total = sum(len(v) for v in blanks.values())
        parts = []
        for c, rows in list(blanks.items())[:8]:
            shown = ", ".join(str(r) for r in rows[:6])
            more = "" if len(rows) <= 6 else f" (+{len(rows) - 6} more)"
            parts.append(f"'{c}' — row(s) {shown}{more}")
        extra = "" if len(blanks) <= 8 else f"; +{len(blanks) - 8} more column(s)"
        msg = (f"Cannot export — {total} blank cell(s) found. Every cell must be "
               f"filled before exporting. Blanks in: " + "; ".join(parts)
               + extra + ".")
        return df, msg

    # --- block export if any cell holds an Excel error literal ---------------
    errs = _excel_error_cells(df)
    if errs:
        total = sum(len(v) for v in errs.values())
        parts = []
        for c, rows in list(errs.items())[:8]:
            shown = ", ".join(str(r) for r in rows[:6])
            more = "" if len(rows) <= 6 else f" (+{len(rows) - 6} more)"
            parts.append(f"'{c}' — row(s) {shown}{more}")
        extra = "" if len(errs) <= 8 else f"; +{len(errs) - 8} more column(s)"
        msg = (f"Cannot export — {total} cell(s) contain an Excel error value "
               f"(e.g. #REF!, #DIV/0!, #N/A). Fix these before exporting: "
               + "; ".join(parts) + extra + ".")
        return df, msg

    # --- fixed-value rules: 'Fix Owner' and 'Verified' -----------------------
    def _rows_not_equal(col: str, expected: str) -> list[int]:
        vals = (df.get_column(col).cast(pl.Utf8, strict=False)
                  .fill_null("").to_list())
        return [i + 1 for i, v in enumerate(vals)
                if str(v).strip().lower() != expected.lower()]

    def _rowlist(rows: list[int]) -> str:
        shown = ", ".join(str(r) for r in rows[:8])
        return shown + ("" if len(rows) <= 8 else f" (+{len(rows) - 8} more)")

    fo_col = _exact_col(df, "Fix Owner")
    if fo_col:
        bad = _rows_not_equal(fo_col, "Vendor")
        if bad:
            return df, (f"Cannot export — 'Fix Owner' must be 'Vendor' on every "
                        f"row. {len(bad)} row(s) are not: row(s) {_rowlist(bad)}.")

    ver_col = _exact_col(df, "Verified")
    if ver_col:
        bad = _rows_not_equal(ver_col, "Pending")
        if bad:
            return df, (f"Cannot export — the 'Verified' column must be 'Pending' "
                        f"on every row. {len(bad)} row(s) are not: "
                        f"row(s) {_rowlist(bad)}.")

    it_col = _exact_col(df, "Issue Type")
    if it_col:
        bad = _rows_not_equal(it_col, "Bug")
        if bad:
            return df, (f"Cannot export — 'Issue Type' must be 'Bug' on every "
                        f"row. {len(bad)} row(s) are not: row(s) {_rowlist(bad)}.")

    # --- Parent must be 'PSGIN-<ID>' and identical on every row --------------
    parent_err = validate_parent_column(df)
    if parent_err:
        return df, "Cannot export — " + parent_err

    return df, None


def validate_parent_column(df: pl.DataFrame) -> str | None:
    """Return a user-facing reason when the 'Parent' column is invalid, else None.

    Rule (Placeholder 3): every row's Parent must be in the 'PSGIN-<ID>' format
    (no spaces; a bare number is rejected) AND every row must carry the SAME
    value. Used both when a delivery sheet is opened and at export time."""
    parent_col = _exact_col(df, "Parent")
    if not parent_col or df.height == 0:
        return None
    vals = (df.get_column(parent_col).cast(pl.Utf8, strict=False)
              .fill_null("").to_list())
    pat = re.compile(r"^PSGIN-\S+$", re.IGNORECASE)
    badfmt = [i + 1 for i, v in enumerate(vals) if not pat.match(str(v).strip())]
    if badfmt:
        shown = ", ".join(str(r) for r in badfmt[:8])
        more = "" if len(badfmt) <= 8 else f" (+{len(badfmt) - 8} more)"
        return ("every 'Parent' value must be in the format 'PSGIN-<ID>' "
                f"(no spaces; a number on its own is not allowed). {len(badfmt)} "
                f"row(s) are not: row(s) {shown}{more}.")
    uniq = sorted({str(v).strip() for v in vals})
    if len(uniq) > 1:
        sample = "; ".join(uniq[:5])
        more = "" if len(uniq) <= 5 else f" (+{len(uniq) - 5} more)"
        return (f"every 'Parent' value must be the same. Found {len(uniq)} "
                f"different values: {sample}{more}.")
    return None


def parent_offending_rows(df: pl.DataFrame) -> list[dict]:
    """Per-row 'Parent' problems for marking + the Issues panel. Returns
    ``[{'index': 0-based, 'row': 1-based, 'value': str, 'problem': str}]`` —
    rows whose Parent isn't 'PSGIN-<ID>' format, or (when all are well-formed)
    rows whose value differs from the majority. Empty when the column is fine."""
    parent_col = _exact_col(df, "Parent")
    if not parent_col or df.height == 0:
        return []
    vals = (df.get_column(parent_col).cast(pl.Utf8, strict=False)
              .fill_null("").to_list())
    pat = re.compile(r"^PSGIN-\S+$", re.IGNORECASE)
    badfmt = [i for i, v in enumerate(vals) if not pat.match(str(v).strip())]
    if badfmt:
        return [{"index": i, "row": i + 1, "value": str(vals[i]).strip(),
                 "problem": "not in 'PSGIN-<ID>' format"} for i in badfmt]
    # All well-formed: flag rows that differ from the majority value.
    from collections import Counter
    counts = Counter(str(v).strip() for v in vals)
    if len(counts) <= 1:
        return []
    majority = counts.most_common(1)[0][0]
    return [{"index": i, "row": i + 1, "value": str(v).strip(),
             "problem": f"differs from the others (expected '{majority}')"}
            for i, v in enumerate(vals) if str(v).strip() != majority]


def export_via_template(df: pl.DataFrame, stem: str = "delivery",
                        title: str = "", course: str = "",
                        details: str = "") -> tuple[Path, bool]:
    """Export a frame using the delivery template if present, else plain xlsx.
    (Placeholder 3 'Export using delivery template'.) Fills Sheet 1's headings
    (title / course / details — never blank) and the derivable dashboard
    numbers, and maps the data into Sheet 2 by matching headers."""
    out = _outpath(stem)
    if WCAG_DELIVERY_TEMPLATE.exists():
        try:
            _fill_delivery(df, out, title=title, course=course, details=details)
            return out, True
        except Exception:  # noqa: BLE001
            log = __import__("logging").getLogger("sde.vpat_report")
            log.exception("delivery template fill failed; falling back to plain")
    from services.excel_service import write_excel, trim_trailing_blank_rows
    write_excel(df, out, sheet_name="Delivery")
    trim_trailing_blank_rows(out)
    return out, False


def _summary_metrics(audit: pl.DataFrame) -> dict:
    """Compute the derivable Sheet-1 dashboard numbers from the loaded audit
    data. Non-derivable metrics (component counts, scan depth) are left out."""
    tags = load_wcag_tags(WCAG_TAGS)
    prio_c = _find_col(audit, ["priority"])
    ref_c = _find_col(audit, ["wcag ref", "wcag sc"])
    owner_c = _find_col(audit, ["fix owner", "owner"])
    rows = audit.to_dicts()
    total = len(rows)

    # Severity buckets follow the upstream Priority labels P1..P4
    # (see dashboard2 SEVERITY_MAP: P1 - Immediate / P2 - High / P3 - Medium /
    # P4 - Low). Legacy Critical/Serious/Moderate/Minor values still map across.
    sev = {"Immediate": 0, "High": 0, "Medium": 0, "Low": 0}
    _pmap = [("p1", "Immediate"), ("immediate", "Immediate"), ("critical", "Immediate"),
             ("p2", "High"), ("high", "High"), ("serious", "High"),
             ("p3", "Medium"), ("medium", "Medium"), ("moderate", "Medium"),
             ("p4", "Low"), ("low", "Low"), ("minor", "Low")]
    n21 = n22 = 0
    fix = {"Direct fix — edit in Moodle": 0, "LMS Admin — template fix": 0,
           "Vendor ticket": 0, "Publisher — eBook source": 0}
    import re as _re
    for r in rows:
        pv = str(r.get(prio_c) or "").lower() if prio_c else ""
        for needle, bucket in _pmap:
            if needle in pv:
                sev[bucket] += 1
                break
        if ref_c:
            m = _re.search(r"(\d\.\d+\.\d+)", str(r.get(ref_c) or ""))
            sc = m.group(1) if m else ""
            if sc in tags:
                level = tags[sc][1]
                lvl = _re.sub(r"\(.*?\)", "", level).strip().upper().replace("LEVEL", "").strip()
                if lvl in ("A", "AA"):
                    if "(added in 2.2)" in level.lower():
                        n22 += 1
                    else:
                        n21 += 1
        if owner_c:
            ov = str(r.get(owner_c) or "").lower()
            if "moodle" in ov or "direct" in ov:
                fix["Direct fix — edit in Moodle"] += 1
            elif "lms" in ov or "template" in ov or "admin" in ov:
                fix["LMS Admin — template fix"] += 1
            elif "vendor" in ov:
                fix["Vendor ticket"] += 1
            elif "publisher" in ov or "ebook" in ov:
                fix["Publisher — eBook source"] += 1
    return {"total": total, "sev": sev, "n21": n21, "n22": n22,
            "best": max(0, total - n21 - n22), "fix": fix}


def _fill_delivery(audit: pl.DataFrame, out: Path, title: str = "",
                   course: str = "", details: str = "") -> None:
    """Fill the delivery template: Sheet 1 (headings + derivable metrics),
    Sheet 2 (data mapped by header), and clear Sheet 3's foreign example data."""
    import datetime as _dt
    from openpyxl import load_workbook
    wb = load_workbook(WCAG_DELIVERY_TEMPLATE)

    # ---- Sheet 2: "All Issues" — map the audit's leading columns POSITIONALLY
    # into however many columns the template's sheet 2 actually has. The audit's
    # columns line up with the delivery columns 1:1 from the left; only column C
    # differs in label (audit "Issue Category" vs template "Issue Type") — same
    # position — so by-name mapping would wrongly pull the audit's separate
    # "Issue Type" (col R) into C. Positional is right.
    #
    # The fill width adapts to the template: it is min(template columns, audit
    # columns), so the sheet works whether the template ships 12 (A–L), 16 (A–P)
    # or any other number of columns — without deleting any template columns.
    issues = next((wb[n] for n in wb.sheetnames
                   if "all issues" in n.lower() or n.strip().startswith("2")), None)
    if issues is not None:
        src_cols = list(audit.columns)            # audit columns, in order
        n_tgt = min(issues.max_column or 0, len(src_cols))
        rows = audit.to_dicts()
        start = 2
        for i, r in enumerate(rows, start=start):
            for j in range(1, n_tgt + 1):
                issues.cell(row=i, column=j, value=r.get(src_cols[j - 1]))
        # clear any leftover template example rows below the data
        for rr in range(start + len(rows), (issues.max_row or 0) + 1):
            for cc in range(1, (issues.max_column or 0) + 1):
                issues.cell(rr, cc).value = None

        # Grid: give every DATA cell a four-side border and strip borders from
        # every empty cell, so the grid bounds exactly the data (no stray
        # template lines below/around it). The header row (row 1) is left exactly
        # as the template ships it. "Has data" uses the same emptiness test as
        # trim_trailing_blank_rows so the two agree.
        from openpyxl.styles import Border, Side
        _thin = Side(style="thin")
        _grid = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        _no_border = Border()
        for rr in range(2, (issues.max_row or 1) + 1):       # skip header row 1
            for cc in range(1, (issues.max_column or n_tgt) + 1):
                c = issues.cell(rr, cc)
                has_data = c.value is not None and str(c.value).strip() != ""
                c.border = _grid if has_data else _no_border
        # The template applies a default border to whole COLUMNS and ROWS (via
        # the column/row dimension styles), so empty cells INHERIT that border
        # and Excel draws a grid down every column / across rows even where there
        # is no data. Per-cell clearing can't reach those (infinite) empty cells,
        # so strip the borderId (StyleArray index 2) from every column/row
        # default style. Data cells keep their own explicit grid; widths/heights
        # are untouched.
        import copy as _copy
        def _strip_border(dim):
            sa = getattr(dim, "_style", None)
            if sa is None:
                return
            try:
                sa = _copy.copy(sa)
                sa[2] = 0                     # borderId -> 0 (no border)
                dim._style = sa
            except Exception:  # noqa: BLE001 — never let cosmetics break export
                pass
        for _dim in list(issues.column_dimensions.values()):
            _strip_border(_dim)
        for _dim in list(issues.row_dimensions.values()):
            _strip_border(_dim)

        # Remove the template's conditional-formatting rule that borders a whole
        # row whenever column A is non-empty (sqref A1:XFD1048576, dxf = thin
        # box). It draws a grid across the ENTIRE row width — including the empty
        # columns past the data (Q onward) — and a conditional format overrides
        # our explicit cell borders, so it must be dropped. Our per-cell grid
        # above already boxes exactly the data cells.
        from openpyxl.formatting.formatting import ConditionalFormattingList
        issues.conditional_formatting = ConditionalFormattingList()

        # Hide Excel's default screen gridlines so the EMPTY cells read as blank
        # (white) and only the real cell borders drawn above are visible — i.e.
        # the grid appears solely where the data is.
        issues.sheet_view.showGridLines = False
        # The template ships a stray saved view (scrolled to col I, cursor parked
        # at J400) that every export inherits and that makes the sheet look wrong
        # on open. Reset it so the sheet opens at the top-left with the header
        # frozen. Reapply the header freeze AFTER clearing topLeftCell so the
        # pane's own anchor is rebuilt cleanly.
        from openpyxl.worksheet.views import Selection
        issues.sheet_view.topLeftCell = None
        issues.sheet_view.selection = [Selection(activeCell="A2", sqref="A2")]
        issues.freeze_panes = "A2"

    # ---- Sheet 1: "Audit Summary" — headings + derivable metrics --------------
    # Layout (delivery template 1.3): a 3-column grid where column B holds the
    # metric label, column C the value and column D the details. The title,
    # footer and section banners are single labels merged across B:D.
    #   B1 (B1:D1)   -> report title
    #   B2 (B2:D2)   -> "COMPLIANCE DASHBOARD" banner
    #   B3/C3        -> header row: Metric | Course Name (C3:D3 merged)
    #   B4 / C4      -> Course                  (value -> C4, C4:D4 merged)
    #   B5 (B5:D6)   -> "AUDIT RESULTS" section header
    #   B7 (B7:D7)   -> "Components in Course Section" banner (no value)
    #   B8..B11      -> Defects Found / WCAG 2.1 AA / WCAG 2.2 AA / Best Practice
    #                   (value -> C, details -> D)
    #   B12 (B12:D13)-> "Standard … | Generated …" footer
    summary = next((wb[n] for n in wb.sheetnames
                    if "audit summary" in n.lower() or n.strip().startswith("1")), None)
    if summary is not None:
        m = _summary_metrics(audit)
        title = (title or "").strip() or "WCAG 2.2 AA Accessibility Audit Summary"
        course = (course or "").strip() or "(not specified)"
        sev = m["sev"]
        # Title (top, B1) and the Standard/Generated footer (bottom, B12).
        summary["B1"] = title
        summary["B12"] = (f"Standard: WCAG22AA | Generated: "
                          f"{_dt.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        # Map by the label in column B; write Value -> C, Details -> D.
        derivable = {
            "defects found": (m["total"],
                              f"Immediate: {sev['Immediate']} | High: {sev['High']} "
                              f"| Medium: {sev['Medium']} | Low: {sev['Low']}"),
            "wcag 2.1 aa issues": (m["n21"],
                                   "Fix first — required for both 2.1 AA and 2.2 AA"),
            "wcag 2.2 aa issues": (m["n22"],
                                   "New in 2.2 only — can defer if targeting 2.1 AA first"),
            "best practice issues": (m["best"], "Not WCAG failures but recommended"),
        }
        # Rows whose label cell (col B) is swallowed by a horizontal merge
        # starting at column A/B (title / footer / section banners) must be left
        # alone — their cells are read-only MergedCells. The header and Course
        # rows merge only from column C (C3:D3, C4:D4), so column B stays its own
        # writable label cell and these rows are still processed below.
        merged_rows = set()
        for rng in summary.merged_cells.ranges:
            if rng.max_col > rng.min_col and rng.min_col <= 2:
                merged_rows.update(range(rng.min_row, rng.max_row + 1))
        for r in range(1, (summary.max_row or 0) + 1):
            if r in merged_rows:
                continue
            label = str(summary.cell(r, 2).value or "").strip().lower()  # column B
            if not label or label in ("—", "metric"):
                continue
            if label == "course":
                # Course name spans the merged C:D cell — write its anchor (C).
                summary.cell(r, 3).value = course    # column C
                continue
            hit = next((v for k, v in derivable.items()
                        if label.startswith(k)), None)
            if hit is not None:
                summary.cell(r, 3).value = hit[0]        # column C
                if hit[1] is not None:
                    summary.cell(r, 4).value = hit[1]    # column D
            else:
                # Non-derivable metric row: clear the template's example
                # value/details so foreign sample data is never shipped.
                summary.cell(r, 3).value = None
                summary.cell(r, 4).value = None

    # ---- Sheet 3: "Component Health" — clear foreign example rows (keep header)
    comp = next((wb[n] for n in wb.sheetnames if "component" in n.lower()), None)
    if comp is not None:
        for rr in range(2, (comp.max_row or 0) + 1):
            for cc in range(1, (comp.max_column or 0) + 1):
                comp.cell(rr, cc).value = None

    wb.save(out)
    from services.excel_service import trim_trailing_blank_rows
    if issues is not None:
        trim_trailing_blank_rows(out, sheet_name=issues.title)
    if comp is not None:
        trim_trailing_blank_rows(out, sheet_name=comp.title)


def summarize_delivery(df: pl.DataFrame) -> dict:
    """Summarise a loaded audit frame by WCAG Ver and Priority (Show errors)."""
    wcag_c = _find_col(df, ["wcag ver", "wcag version", "wcag sc", "wcag",
                            "success criteria", "criterion", "criteria"])
    prio_c = _find_col(df, ["priority", "prio", "urgency", "severity"])
    out = {"total": df.height, "wcag_col": wcag_c, "prio_col": prio_c,
           "by_wcag": {}, "by_prio": {}}

    def _counts(col):
        g = (df.with_columns(pl.col(col).cast(pl.Utf8, strict=False).fill_null(""))
             .filter(pl.col(col).str.strip_chars() != "")
             .group_by(col).len().sort("len", descending=True))
        return {r[col]: r["len"] for r in g.to_dicts()}

    if wcag_c:
        out["by_wcag"] = _counts(wcag_c)
    if prio_c:
        out["by_prio"] = _counts(prio_c)
    return out


def _fill_template(audit: pl.DataFrame, out: Path, template: Path = WCAG_TEMPLATE) -> None:
    """Fill a template's '2 - All Issues' sheet, writing data right after the
    header row (overwriting any pre-formatted blank rows), then drop leftover
    trailing blank rows on that sheet so there is no gap and no empty tail."""
    from openpyxl import load_workbook
    wb = load_workbook(template)
    ws = None
    for name in wb.sheetnames:
        if "all issues" in name.lower() or name.strip().startswith("2"):
            ws = wb[name]
            break
    ws = ws or wb.active
    # Header = first non-empty row (normally row 1); data begins immediately after.
    header_row = 1
    for r in range(1, (ws.max_row or 1) + 1):
        if any(ws.cell(r, c).value not in (None, "")
               for c in range(1, (ws.max_column or 1) + 1)):
            header_row = r
            break
    start = header_row + 1
    for i, row in enumerate(audit.iter_rows(), start=start):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    wb.save(out)
    # Remove any pre-formatted blank rows left below the data on the issues sheet.
    from services.excel_service import trim_trailing_blank_rows
    trim_trailing_blank_rows(out, sheet_name=ws.title)


def vpat_report(path: str | Path) -> tuple[list, dict, dict]:
    df = _read_table(path)
    if df.height == 0:
        raise ValueError("The uploaded audit workbook has no data rows.")

    desc_c = _find_col(df, ["issue description", "description", "issue", "help",
                            "message", "summary", "violation", "finding"])
    wcag_c = _find_col(df, ["wcag sc", "wcag", "success criteria", "criterion",
                            "criteria", "guideline", "standard", "sc"])
    impact_c = _find_col(df, ["impact", "severity", "priority", "level"])
    prio_c = _find_col(df, ["priority", "urgency", "p1", "rank"])

    if desc_c is None and wcag_c is None:
        raise ValueError(
            "Could not recognise audit columns. Expected an issue/description "
            "column and ideally a WCAG criterion column. Found: "
            + ", ".join(df.columns[:12]) + ("…" if df.width > 12 else "")
        )

    work = df.select([
        _col_or_blank(df, desc_c).alias("Issue Description"),
        _col_or_blank(df, wcag_c).alias("WCAG SC"),
        _col_or_blank(df, impact_c).alias("Severity"),
        _col_or_blank(df, prio_c).alias("Priority"),
    ]).filter(
        (pl.col("Issue Description").str.strip_chars() != "")
        | (pl.col("WCAG SC").str.strip_chars() != "")
    )
    total = work.height
    if total == 0:
        raise ValueError("No accessibility findings were detected in the workbook.")

    unique_issues = work.unique(subset=["WCAG SC", "Issue Description"],
                                maintain_order=True)
    n_unique = unique_issues.height

    def _breakdown(col: str) -> dict:
        g = (work.filter(pl.col(col).str.strip_chars() != "")
             .group_by(col).len().sort("len", descending=True))
        return {row[col]: row["len"] for row in g.to_dicts()}

    by_wcag = _breakdown("WCAG SC")
    by_sev = _breakdown("Severity")
    by_prio = _breakdown("Priority")

    # VPAT conformance table (unique issues grouped by criterion)
    crit = (unique_issues.filter(pl.col("WCAG SC").str.strip_chars() != "")
            .group_by("WCAG SC")
            .agg(pl.len().alias("count"),
                 pl.col("Issue Description").first().alias("sample"))
            .sort("count", descending=True))
    vpat_rows = []
    for row in crit.to_dicts():
        n = row["count"]
        vpat_rows.append({
            "Criteria": row["WCAG SC"],
            "Conformance Level": "Does Not Support" if n >= 3 else "Partially Supports",
            "Issue Count": n,
            "Remarks and Explanations":
                f"{n} issue(s) identified. e.g. {str(row['sample'])[:160]}",
        })
    if not vpat_rows:
        for k, v in (by_sev or by_prio).items():
            vpat_rows.append({"Criteria": k, "Conformance Level": "Partially Supports",
                              "Issue Count": v,
                              "Remarks and Explanations": f"{v} issue(s)."})

    # build outputs: findings workbook (xlsx) + csv, plus the VPAT report
    from services.excel_service import write_excel
    findings = work.with_columns(pl.Series("S.No", list(range(1, total + 1)))) \
                   .select(["S.No", "Issue Description", "WCAG SC", "Severity", "Priority"])
    xlsx_out = _outpath("vpat_findings")
    write_excel(findings, xlsx_out, sheet_name="Findings")
    csv_out = _outpath("vpat_findings", ".csv")
    findings.write_csv(csv_out)
    # 3) VPAT report workbook
    summary_rows = [
        {"Metric": "Total findings", "Value": total},
        {"Metric": "Unique issues", "Value": n_unique},
        {"Metric": "WCAG criteria affected", "Value": len(by_wcag)},
    ] + [{"Metric": f"Severity – {k}", "Value": v} for k, v in by_sev.items()] \
      + [{"Metric": f"Priority – {k}", "Value": v} for k, v in by_prio.items()]
    vpat_out = _outpath("vpat_report")
    _write_vpat_workbook(vpat_out, summary_rows, vpat_rows)

    stats = {
        "total_findings": total,
        "unique_issues": n_unique,
        "criteria_affected": len(by_wcag),
        "by_wcag": by_wcag,
        "by_severity": by_sev,
        "by_priority": by_prio,
    }
    outputs = [("Delivery file (Excel)", xlsx_out),
               ("Delivery file (CSV)", csv_out),
               ("Summary report", vpat_out)]
    return outputs, stats, _preview(findings)


# ---------------------------------------------------------------------------
# Import Digital Asset (Placeholder 3)
#
# A delivery audit is already open in the editor. The user picks a SEPARATE
# "downloadables" workbook; its 'Counts by Type' sheet (column B, fixed rows
# 7 = PDF, 8 = PowerPoint, 9 = Word) gives the three document-type counts.
#
# THREE rows from Digital_Asset_Template.xlsx (Word / PDF / PowerPoint) are then
# appended to the END of the loaded dataset, copied verbatim from the template
# except for the fields that must fit the loaded audit and Placeholder-3 rules:
#   * Instances -> that type's count from 'Counts by Type'
#   * ID        -> continues the loaded dataset's 'Issue - NN' sequence
#   * Summary   -> regenerated so its leading number matches the new ID
#   * Parent    -> the loaded dataset's Parent (so every Parent stays identical)
#   * Fix Owner -> 'Vendor'   and   Verified -> 'Pending'   (placeholder-3 rules)
# Placeholder 4's classifier is reused only to map each template row to its type.
# ---------------------------------------------------------------------------


def _read_digital_asset_template() -> dict[str, dict]:
    """Read the boilerplate Digital_Asset_Template.xlsx into
    ``{friendly_type: {audit_header: value}}``. Empty cells become "".

    Each row's friendly type (PDF / Microsoft Word / Microsoft PowerPoint) is
    derived with the shared classifier from the row's own text, so the mapping
    is never hardcoded."""
    if not DIGITAL_ASSET_TEMPLATE.exists():
        raise ValueError(
            "The Digital Asset template is missing on the server "
            f"({DIGITAL_ASSET_TEMPLATE.name}). Sync it into the templates "
            "volume and try again.")
    df = _read_table(DIGITAL_ASSET_TEMPLATE)
    out: dict[str, dict] = {}
    for r in df.to_dicts():
        friendly = _classify_doc(
            str(r.get("Issue Category") or ""), str(r.get("Element") or ""),
            str(r.get("Component") or ""), str(r.get("Issue") or ""))
        if friendly and friendly not in out:
            out[friendly] = {k: ("" if v is None else str(v)) for k, v in r.items()}
    if not out:
        raise ValueError("The Digital Asset template has no recognisable rows.")
    return out


# 'Counts by Type' sheet layout: column A = Media type, column B = Count, with
# the three document types at fixed rows: 7 = PDF, 8 = PowerPoint, 9 = Word.
COUNTS_SHEET_NAME = "Counts by Type"
_COUNTS_ROWS = {"PDF": 7, "Microsoft PowerPoint": 8, "Microsoft Word": 9}
_COUNTS_COL = 2  # column B holds the count


def _find_counts_sheet(wb):
    """Return the worksheet whose name is (or resembles) 'Counts by Type',
    else None."""
    for name in wb.sheetnames:
        if name.strip().lower() == COUNTS_SHEET_NAME.lower():
            return wb[name]
    for name in wb.sheetnames:  # tolerant fallback: any "...count...type..." sheet
        n = name.lower()
        if "count" in n and "type" in n:
            return wb[name]
    return None


def read_counts_by_type(path: str | Path) -> dict | None:
    """Read PDF / PowerPoint / Word counts from the loaded workbook's
    'Counts by Type' sheet (column B at the fixed rows 7 / 8 / 9).

    Returns ``{"PDF": n, "Microsoft PowerPoint": n, "Microsoft Word": n}`` or
    ``None`` when the workbook has no 'Counts by Type' sheet. Non-numeric or
    negative cells are treated as 0."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = _find_counts_sheet(wb)
        if ws is None:
            return None

        def _num(v) -> int:
            s = str(v if v is not None else "").strip()
            if not s:
                return 0
            try:
                return max(0, int(float(s)))
            except (ValueError, TypeError):
                return 0

        return {friendly: _num(ws.cell(row=r, column=_COUNTS_COL).value)
                for friendly, r in _COUNTS_ROWS.items()}
    finally:
        wb.close()


def next_issue_seq(df: pl.DataFrame) -> int:
    """Next 'Issue - NN' number, continuing from the largest trailing number in
    the dataset's ID column (1 when none/empty). Matches the numbering check's
    trailing-digit rule so the appended IDs stay contiguous."""
    id_c = _exact_col(df, "ID") or (df.columns[0] if df.columns else None)
    if not id_c or df.height == 0:
        return 1
    nums = (df.get_column(id_c).cast(pl.Utf8, strict=False)
              .str.extract(r"(\d+)\s*$", 1)
              .cast(pl.Int64, strict=False).drop_nulls().to_list())
    return (max(nums) + 1) if nums else 1


def most_common_parent(df: pl.DataFrame) -> str:
    """The most frequent non-blank value in the dataset's Parent column (so the
    appended rows inherit the same Parent). "" when none."""
    pc = _exact_col(df, "Parent")
    if not pc or df.height == 0:
        return ""
    g = (df.with_columns(pl.col(pc).cast(pl.Utf8, strict=False).fill_null(""))
           .filter(pl.col(pc).str.strip_chars() != "")
           .group_by(pc).len().sort("len", descending=True))
    rows = g.to_dicts()
    return str(rows[0][pc]).strip() if rows else ""


# Per-type label used in the Issue text count suffix.
_COUNT_LABEL = {"PDF": "PDFs", "Microsoft Word": "Word documents",
                "Microsoft PowerPoint": "PowerPoint documents"}

# Order of the labelled sections that make up a digital-asset 'Description',
# mapping each section heading to the audit column it is built from.
_DESC_SECTIONS = [
    ("Issue Type", "Issue Category"), ("Root Cause", "Root Cause"),
    ("Issue", "Issue"), ("Element", "Element"), ("Resolution", "Resolution"),
    ("Fix Owner", "Fix Owner"), ("WCAG Ref", "WCAG Ref"), ("WCAG", "WCAG"),
]


def _issue_with_count(issue: str, friendly: str, count: int) -> str:
    """Append the document count to the Issue text, e.g.
    'Ensure downloaded PDF documents are accessible … technologies'
    -> '… technologies. Number of PDFs: 354'."""
    core = re.sub(r"\s*\.?\s*$", "", (issue or "").strip())   # drop trailing '.'
    label = _COUNT_LABEL.get(friendly, "documents")
    return f"{core}. Number of {label}: {count}"


def _build_description(values: dict) -> str:
    """Build a digital-asset 'Description' in the template's labelled-section
    format from the row's FINAL values, so it reflects the updated Issue text,
    Fix Owner, etc. The 'WCAG' section is rendered as '<SC> <name>'
    (e.g. '1.3.1 Info and Relationships') to match the other rows' format —
    the SC is taken from 'WCAG Ref'."""
    out = dict(values)
    m = re.search(r"(\d\.\d+\.\d+)", str(values.get("WCAG Ref", "")))
    name = str(values.get("WCAG", "")).strip()
    out["WCAG"] = f"{m.group(1)} {name}".strip() if m else name
    body = "\n\n".join(f"{label}:\n{out.get(col, '')}"
                       for label, col in _DESC_SECTIONS)
    return "HOW TO FIND:\n" + body


def build_digital_asset_rows(counts: dict, existing_df: pl.DataFrame,
                             parent_id: str = "", start_seq: int = 1,
                             issue_url: str = ""):
    """Build the three digital-asset rows to append to the loaded dataset.

    Each boilerplate row from Digital_Asset_Template.xlsx (template order:
    Word / PDF / PowerPoint) is copied VERBATIM except for the fields that must
    fit the loaded audit + Placeholder-3 rules:
      Instances -> that type's count from ``counts``
      ID        -> 'Issue - NN' continuing from ``start_seq``
      Summary   -> regenerated so its leading number matches the new ID
      Parent    -> ``parent_id`` (kept identical across rows)
      Issue URL -> ``issue_url`` (the URL the user typed; same on all 3 rows)
      Fix Owner -> 'Vendor'   and   Verified -> 'Pending'

    ``existing_df`` supplies the column order so the result lines up for concat
    (falls back to the 23 audit headers when None).

    Returns ``(new_rows_df, by)`` where ``new_rows_df`` is an all-Utf8 Polars
    frame (always 3 rows) and ``by`` is ``{"pdf": n, "word": n, "ppt": n}``."""
    headers = list(existing_df.columns) if existing_df is not None else audit_headers()
    template = _read_digital_asset_template()  # insertion order == template order

    rows: list[dict] = []
    seq = int(start_seq)
    for friendly, base in template.items():
        cnt = int(counts.get(friendly, 0) or 0)
        if cnt <= 0:
            continue                           # skip types with a 0 count
        new_id = f"Issue - {seq:02d}"
        seq += 1
        row = dict(base)                       # every column verbatim from template
        # WCAG Ref: only the criterion number, no 'Ref:' prefix.
        _refm = re.search(r"(\d\.\d+\.\d+)", str(base.get("WCAG Ref", "")))
        row["ID"] = new_id
        row["Priority"] = "P2 - High"          # column B is always 'P2 - High'
        row["Instances"] = str(cnt)
        row["Pages"] = "0"                      # 'Pages' is always 0
        row["Issue"] = _issue_with_count(str(base.get("Issue", "")), friendly, cnt)
        row["WCAG Ver"] = "WCAG2A"             # fixed conformance tag
        row["WCAG Ref"] = _refm.group(1) if _refm else str(base.get("WCAG Ref", "")).strip()
        # Resolution: drop the leading 'HOW TO FIX -' line.
        row["Resolution"] = re.sub(r"(?i)^[ \t]*HOW TO FIX[^\n]*\r?\n?", "",
                                   str(base.get("Resolution", "")))
        row["Issue URL"] = (issue_url or "").strip()
        row["Parent"] = parent_id or str(base.get("Parent") or "")
        row["Fix Owner"] = "Vendor"
        row["Verified"] = "Pending"
        row["Issue Type"] = "Bug"              # column R is 'Bug' (not 'Story')
        name = str(base.get("Component") or "").strip()
        row["Summary"] = f'{new_id} - {row.get("Issue Category", "")} - {name} - {row["Issue"]}'
        row["Description"] = _build_description(row)   # rebuilt from final values
        rows.append({h: row.get(h, "") for h in headers})

    by = {"pdf": int(counts.get("PDF", 0) or 0),
          "word": int(counts.get("Microsoft Word", 0) or 0),
          "ppt": int(counts.get("Microsoft PowerPoint", 0) or 0)}
    new_df = pl.DataFrame(rows, schema={h: pl.Utf8 for h in headers})
    return new_df, by


# ---------------------------------------------------------------------------
# Import Alt Text (Placeholder 3)
#
# Appends ONE fixed audit row (the "Audio Description" alt-text boilerplate from
# Alt_Text_Template.xlsx) to the end of the loaded dataset. Only the dynamic
# fields are changed: ID continues the 'Issue - NN' sequence, Parent is the
# dataset's value, and the Summary is regenerated so its leading number matches
# the new ID. The Description is spaced like the loaded rows. Everything else is
# copied verbatim from the template.
# ---------------------------------------------------------------------------

def _read_alt_text_template() -> dict:
    """Read the single boilerplate row from Alt_Text_Template.xlsx as
    {audit_header: value}. Empty cells become ""."""
    if not ALT_TEXT_TEMPLATE.exists():
        raise ValueError(
            "The Alt Text template is missing on the server "
            f"({ALT_TEXT_TEMPLATE.name}). Sync it into the templates volume "
            "and try again.")
    df = _read_table(ALT_TEXT_TEMPLATE)
    if df.height == 0:
        raise ValueError("The Alt Text template has no data row.")
    r = df.to_dicts()[0]
    return {k: ("" if v is None else str(v)) for k, v in r.items()}


def build_alt_text_rows(existing_df: pl.DataFrame, parent_id: str = "",
                        start_seq: int = 1):
    """Build the single alt-text row to append.

    Copies the template verbatim except: ID -> 'Issue - NN' (from ``start_seq``),
    Parent -> ``parent_id``, Summary -> regenerated with the new ID, and the
    Description spaced for readability. ``existing_df`` supplies the column order
    (falls back to the 23 audit headers when None). Returns a 1-row Polars
    frame."""
    headers = list(existing_df.columns) if existing_df is not None else audit_headers()
    base = _read_alt_text_template()
    new_id = f"Issue - {int(start_seq):02d}"
    row = dict(base)
    row["ID"] = new_id
    row["Parent"] = parent_id or str(base.get("Parent") or "")
    comp = str(base.get("Component") or "").split("?")[0].strip()
    row["Summary"] = (f'{new_id} - {base.get("Issue Category", "")} - '
                      f'{comp} - {base.get("Issue", "")}')
    # Rebuild the Description from the row's columns so it always stays
    # consistent with them (and validates clean).
    row["Description"] = _build_description(row)
    new_df = pl.DataFrame([{h: row.get(h, "") for h in headers}],
                          schema={h: pl.Utf8 for h in headers})
    return new_df


def _write_vpat_workbook(out: Path, summary_rows: list[dict],
                         vpat_rows: list[dict]) -> Path:
    import xlsxwriter
    with xlsxwriter.Workbook(str(out)) as wb:
        title = wb.add_format({"bold": True, "font_size": 14, "font_color": "#1d4ed8"})
        hdr = wb.add_format({"bold": True, "border": 1})
        cell = wb.add_format({"border": 1, "valign": "top", "text_wrap": True})
        s = wb.add_worksheet("Summary")
        s.merge_range(0, 0, 0, 1, "Product Accessibility Report — Summary", title)
        s.write(2, 0, "Metric", hdr); s.write(2, 1, "Value", hdr)
        for i, row in enumerate(summary_rows, start=3):
            s.write(i, 0, row["Metric"], cell); s.write(i, 1, row["Value"], cell)
        s.set_column(0, 0, 34); s.set_column(1, 1, 16)
        v = wb.add_worksheet("VPAT")
        cols = ["Criteria", "Conformance Level", "Issue Count", "Remarks and Explanations"]
        v.merge_range(0, 0, 0, len(cols) - 1,
                      "WCAG 2.x Report (auto-generated — review before publishing)", title)
        for j, c in enumerate(cols):
            v.write(2, j, c, hdr)
        for i, row in enumerate(vpat_rows, start=3):
            for j, c in enumerate(cols):
                v.write(i, j, row.get(c, ""), cell)
        v.set_column(0, 0, 26); v.set_column(1, 1, 20)
        v.set_column(2, 2, 12); v.set_column(3, 3, 70)
        v.freeze_panes(3, 0)
    return out
