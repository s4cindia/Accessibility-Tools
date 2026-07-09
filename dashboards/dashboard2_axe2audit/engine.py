"""Placeholder 2 (axe -> audit conversion) engine.

This placeholder's own working engine. Shared low-level primitives
(Excel read/write, WCAG tag-file, audit header template) are imported
from services.feature_service so they are defined once.
"""
from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

import polars as pl

from config import config
from services import export_service as _ex
from services.excel_service import normalize_xlsx_text, read_sheet

from services.feature_service import (
    TEMPLATE_DIR, WCAG_TEMPLATE, WCAG_DELIVERY_TEMPLATE, WCAG_TAGS,
    list_sheets_for, template_status, audit_headers, validate_audit_columns,
    fill_issue_category, strip_description_hyphen, load_wcag_tags,
    EXCEL_ERROR_VALUES, _read_table, _find_col, _col_or_blank, _outpath, _ts,
    _slug, _preview, _exact_col, _category_from_issue, _excel_error_cells,
    _AUDIT_HEADERS_FALLBACK,
)


TARGET_SHEET = "2 - All Issues"


DEFAULT_INPUT_START_ROW = 2


DEFAULT_OUTPUT_START_ROW = 2


COLUMN_MAP = [("B", "D"), ("O", "E"), ("A", "G"), ("R", "I"), ("J", "J")]


SEVERITY_COLUMN = "E"


SEVERITY_PRIORITY_COLUMN = "B"


SEVERITY_URGENCY_COLUMN = "U"


SEVERITY_MAP = {
    "critical": ("P1 - Immediate", "1 - Urgent"),
    "serious":  ("P2 - High",      "2 - High"),
    "moderate": ("P3 - Medium",    "3 - Medium"),
    "minor":    ("P4 - Low",       "4 - Low"),
}


WCAG_TAG_COLUMN = "L"


WCAG_ID_COLUMN = "N"


WCAG_NAME_COLUMN = "O"


WCAG_LEVEL_COLUMN = "F"


ISSUE_DESC_COLUMN = "B"


ISSUE_CATEGORY_COLUMN = "C"


_WCAG_TAG_RE = re.compile(r"^wcag(\d{3,4})([a-z]?)$", re.IGNORECASE)


_WCAGAA_TAG_RE = re.compile(r"^wcag(\d{1,2}a{1,3})$", re.IGNORECASE)

# Parent ID is written to every output row as "PSGIN-<id>" (the project key
# "PSGIN" — see column Q — joined to the supplied Parent ID with no spaces).
PARENT_PREFIX = "PSGIN"


def _format_parent_id(parent_id: str) -> str:
    """Return the Parent ID in the 'PSGIN-<id>' format (no spaces). A leading
    'PSGIN' prefix the user already typed (e.g. from the prefilled input) is
    stripped first, so it's never doubled and a prefix-only value yields ''."""
    pid = (parent_id or "").strip()
    if not pid:
        return ""
    core = re.sub(rf"^{PARENT_PREFIX}\s*-?\s*", "", pid, flags=re.IGNORECASE).strip()
    return f"{PARENT_PREFIX}-{core}" if core else ""


_STANDALONE_HEADERS = {
    "A": "Issue ID", "B": "Priority", "C": "Issue Category", "D": "Description",
    "E": "Component / Rule", "F": "Level / Standard", "G": "Rule ID",
    "H": "Element / Selector", "I": "URL / Screenshot", "J": "Resolution",
    "K": "Fix Owner", "L": "Reopen Count", "M": "Comment Count",
    "N": "WCAG SC", "O": "WCAG Name", "P": "Status",
    "Q": "Project", "R": "Type", "S": "Summary", "T": "Details", "U": "Urgency",
    "V": "Owner Email", "W": "Parent ID",
}


_STANDALONE_WIDTHS = {
    "A": 12, "B": 15, "C": 22, "D": 48, "E": 20, "F": 16, "G": 18, "H": 36,
    "I": 30, "J": 40, "K": 12, "L": 12, "M": 13, "N": 10, "O": 26, "P": 12,
    "Q": 10, "R": 12, "S": 50, "T": 64, "U": 14, "V": 30, "W": 14,
}


_STANDALONE_WRAP = {"C", "D", "H", "I", "J", "O", "S", "T"}


_KEYWORD_RULES = [
    ("contrast", "Color Contrast"),
    ("focus indicator", "CSS (Focus)"), ("focus visible", "CSS (Focus)"),
    ("focus-visible", "CSS (Focus)"), ("focus outline", "CSS (Focus)"),
    ("keyboard focus", "CSS (Focus)"), (":focus", "CSS (Focus)"),
    ("keyboard alone", "Keyboard"), ("keyboard operable", "Keyboard"),
    ("keyboard accessible", "Keyboard"), ("by keyboard", "Keyboard"),
    ("keyboard", "Keyboard"),
    ("status message", "Status Messages"),
    ("automatically announced", "Status Messages"),
    ("live region", "Status Messages"),
    ("landmark", "Landmarks"), ("region", "Landmarks"),
    ("heading", "Headings"),
    ("list item", "Lists"), ("<li>", "Lists"),
    ("iframe", "Frames / Iframes"), ("frame", "Frames / Iframes"),
    ("label", "Forms / Labels"), ("form field", "Forms / Labels"),
    ("input field", "Forms / Labels"), ("autocomplete", "Forms / Labels"),
    ("touch target", "Target Size"), ("target size", "Target Size"),
    ("minimum size", "Target Size"), ("size or spacing", "Target Size"),
    ("size and space", "Target Size"),
    ("alt text", "Image"), ("alternative text", "Image"),
    ("text alternative", "Image"), ("alt attribute", "Image"),
    ("convey", "Use of Color / State Indication"),
    ("color alone", "Use of Color / State Indication"),
    ("button", "Buttons"),
    ("accessible name", "ARIA / Widget"), ("aria", "ARIA / Widget"),
    ("role", "ARIA / Widget"), ("interactive control", "ARIA / Widget"),
]


_KEYWORD_RULES += [
    # --- time-based media: SPECIFIC first ---
    ("audio description",   "Audio Description"),      # 1.2.1 / 1.2.3  (must precede "audio")
    ("video-only",          "Audio Description"),      # 1.2.1
    ("transcript",          "Media Alternative"),      # 1.2.1 / 1.2.8
    ("subtitle",            "Captions"),               # precede "title" (substring!)
    ("caption",             "Captions"),               # 1.2.2
    ("audio",               "Audio Control"),          # 1.4.2

    # --- moving content: "blinking" MUST precede "link" (substring!) ---
    ("animation",           "Pause / Stop / Hide"),    # 2.2.2
    ("blinking",            "Pause / Stop / Hide"),
    ("scrolling",           "Pause / Stop / Hide"),
    ("moving",              "Pause / Stop / Hide"),

    # --- images ---
    ("decorative",          "Image"),                  # 1.1.1
    ("image",               "Image"),

    # --- reflow / resize / spacing ---
    ("reflow",              "Reflow"),                  # 1.4.10
    ("320px",               "Reflow"),
    ("width equivalent",    "Reflow"),
    ("text spacing",        "Text Spacing"),           # 1.4.12
    ("200%",                "Resize Text"),             # 1.4.4
    ("zoom",                "Resize Text"),
    ("text scaling",        "Resize Text"),
    ("viewport",            "Resize Text"),

    # --- structure / metadata ---
    ("orientation",         "Orientation"),            # 1.3.4
    ("visual boundary",     "Non-text Contrast"),      # 1.4.11
    ("lang attribute",      "Language of Page"),       # 3.1.1
    ("subtitle",            "Captions"),               # (dup guard, harmless)
    ("title",               "Page Titled"),            # 2.4.2  (after subtitle)
    ("modal",               "Keyboard"),               # 2.1.2  (keyboard trap / ESC)
    ("esc key",             "Keyboard"),
    ("link",                "Links"),                  # 2.4.4  (after "blinking")
    ("list",                "Lists"),                  # 1.3.1  (after "list item"/"<li>")
]


_RULEID_CATEGORY_HINTS = {
    "color-contrast": {"Color Contrast"},
    "link-in-text-block": {"Color Contrast", "Use of Color / State Indication", "Links"},
    "focus": {"CSS (Focus)", "Keyboard"},
    "landmark": {"Landmarks"}, "region": {"Landmarks"},
    "heading": {"Headings"}, "empty-heading": {"Headings"},
    "page-has-heading-one": {"Headings"},
    "listitem": {"Lists"}, "definition-list": {"Lists"}, "list": {"Lists"},
    "frame-title": {"Frames / Iframes"}, "frame-tested": {"Frames / Iframes"},
    "frame": {"Frames / Iframes"},
    "label-title-only": {"Forms / Labels"},
    "label-content-name-mismatch": {"Forms / Labels", "ARIA / Widget"},
    "form-field-multiple-labels": {"Forms / Labels"},
    "select-name": {"Forms / Labels"}, "autocomplete-valid": {"Forms / Labels"},
    "label": {"Forms / Labels"},
    "image-alt": {"Image"}, "input-image-alt": {"Image"},
    "role-img-alt": {"Image"}, "svg-img-alt": {"Image"}, "area-alt": {"Image"},
    "image-redundant-alt": {"Image"}, "object-alt": {"Image"},
    "button-name": {"Buttons", "ARIA / Widget"},
    "input-button-name": {"Buttons", "ARIA / Widget"},
    "link-name": {"Buttons", "ARIA / Widget", "Links"},
    "target-size": {"Target Size"},
    "aria-live": {"Status Messages", "ARIA / Widget"},
    "nested-interactive": {"Keyboard", "ARIA / Widget"},
    "scrollable-region-focusable": {"Keyboard"}, "keyboard": {"Keyboard"},
    "aria-": {"ARIA / Widget", "Forms / Labels", "Status Messages", "Buttons", "Image"},
    "aria": {"ARIA / Widget"}, "role": {"ARIA / Widget"},
}


def _wcag_ids_from_cell(value) -> list[str]:
    if value is None:
        return []
    seen = []
    for token in str(value).split(","):
        m = _WCAG_TAG_RE.match(token.strip().lower())
        if m:
            digits, suffix = m.group(1), m.group(2)
            criterion = f"{digits[0]}.{digits[1]}.{digits[2:]}"
            if suffix:
                criterion += "." + suffix
            if criterion not in seen:
                seen.append(criterion)
    return seen


def _category_allowed_for_ruleid(category, ruleid) -> bool:
    if not ruleid:
        return True
    rid = str(ruleid).strip().lower()
    if not rid:
        return True
    for key in sorted(_RULEID_CATEGORY_HINTS, key=len, reverse=True):
        if key in rid:
            return category in _RULEID_CATEGORY_HINTS[key]
    return True


def _category_from_keywords(description, ruleid=None):
    if not description:
        return None
    text = str(description).lower()
    category = None
    for needle, cat in _KEYWORD_RULES:
        if needle in text:
            category = cat
            break
    if category is None:
        return None
    if not _category_allowed_for_ruleid(category, ruleid):
        return None
    return category


def _input_worksheet(path: Path, sheet: str | None):
    """Return (worksheet, max_row, close_fn). Supports xlsx/xls and csv/tsv."""
    from openpyxl import Workbook, load_workbook
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        import io
        from services.csv_service import decode_csv_bytes
        sep = "\t" if suffix == ".tsv" else ","
        # Tolerate cp1252/latin-1 CSVs (Excel/Windows exports) — a plain
        # pl.read_csv on those raises "invalid utf-8 sequence".
        df = pl.read_csv(io.BytesIO(decode_csv_bytes(path)), separator=sep,
                         infer_schema_length=0, truncate_ragged_lines=True,
                         ignore_errors=True)
        wb = Workbook()
        ws = wb.active
        ws.append(list(df.columns))
        for row in df.iter_rows():
            ws.append(list(row))
        return ws, ws.max_row, wb.close
    wb = load_workbook(path, data_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "Merged" in wb.sheetnames:
        ws = wb["Merged"]
    else:
        ws = wb.active
    return ws, ws.max_row, wb.close


_RULE_CATALOGUE_CACHE = None


def _rule_catalogue() -> dict:
    """Map every rule id in Rule_ID_Mapping.xlsx to its Tags string (cached).
    Used as a fallback to fill WCAG data when an input row carries no tag."""
    global _RULE_CATALOGUE_CACHE
    if _RULE_CATALOGUE_CACHE is not None:
        return _RULE_CATALOGUE_CACHE
    cat = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(TEMPLATE_DIR / "Rule_ID_Mapping.xlsx",
                           read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        if rows:
            hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
            ridx = hdr.index("rule id") if "rule id" in hdr else 0
            tidx = hdr.index("tags") if "tags" in hdr else None
            for r in rows[1:]:
                if r and len(r) > ridx and r[ridx]:
                    tags = (str(r[tidx]).strip()
                            if tidx is not None and len(r) > tidx and r[tidx] else "")
                    cat[str(r[ridx]).strip()] = tags
    except Exception:  # noqa: BLE001 - a missing/locked catalogue is non-fatal
        pass
    _RULE_CATALOGUE_CACHE = cat
    return cat


def axe_to_audit(path: str | Path, sheet: str | None = None,
                 parent_id: str = "", out_name: str = "") -> tuple[list, dict, dict]:
    """Faithful web port of generate_axe_excel.fill_template()."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    path = Path(path)
    in_ws, in_max, close_src = _input_worksheet(path, sheet)

    tags_found = WCAG_TAGS.exists()
    tags_map = load_wcag_tags(WCAG_TAGS)

    # output workbook: copy of the template, or a standalone fallback
    template_used = False
    if WCAG_TEMPLATE.exists():
        tpl = load_workbook(WCAG_TEMPLATE)
        if TARGET_SHEET in tpl.sheetnames:
            out_ws = tpl[TARGET_SHEET]
        else:
            out_ws = tpl.active
        out_wb = tpl
        template_used = True
    else:
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = TARGET_SHEET
        for col, label in _STANDALONE_HEADERS.items():
            out_ws[f"{col}1"] = label

    in_start = DEFAULT_INPUT_START_ROW
    out_row = DEFAULT_OUTPUT_START_ROW
    written = wcag_rows = matched = severity_rows = 0
    missing: set[str] = set()

    def _cv(col, row):
        return in_ws[f"{col}{row}"].value

    for in_row in range(in_start, in_max + 1):
        cells = {}
        first_id = name = level = ""
        non_empty = False

        # 1. straight column copies
        for in_col, out_col in COLUMN_MAP:
            v = _cv(in_col, in_row)
            cells[out_col] = v
            if v is not None and str(v).strip() != "":
                non_empty = True

        # 2/3. WCAG id (first only) + name/level lookup
        ids = _wcag_ids_from_cell(_cv(WCAG_TAG_COLUMN, in_row))
        if ids:
            first_id = ids[0]
            # Trim a stray trailing letter ('1.1.1.b' -> '1.1.1') so the id
            # matches the tag map and the WCAG columns are never left blank.
            first_id = cells[WCAG_ID_COLUMN] = (
                first_id.rsplit(".", 1)[0]
                if first_id and first_id[-1].isalpha() else first_id)
            non_empty = True
            wcag_rows += 1
            if first_id in tags_map:
                name, level = tags_map[first_id]
                cells[WCAG_NAME_COLUMN] = f"{name}"
                cells[WCAG_LEVEL_COLUMN] = level
                matched += 1
            else:
                missing.add(first_id)

        # 4. severity -> priority / urgency
        sev_raw = _cv(SEVERITY_COLUMN, in_row)
        if sev_raw is not None:
            sev = str(sev_raw).strip().lower()
            if sev in SEVERITY_MAP:
                priority, urgency = SEVERITY_MAP[sev]
                cells[SEVERITY_PRIORITY_COLUMN] = priority
                cells[SEVERITY_URGENCY_COLUMN] = urgency
                non_empty = True
                severity_rows += 1

        # 4b. issue category from description, validated against rule id (A)
        desc_val = _cv(ISSUE_DESC_COLUMN, in_row)
        ruleid_val = _cv("A", in_row)
        issue_category = _category_from_keywords(desc_val, ruleid_val)
        # Conversion condition (moved here from placeholder 3): if no category
        # could be derived from keywords, fall back to the Issue (rule id, G) —
        # hyphens -> spaces, Title Case — so Issue Category is never left blank.
        if not issue_category:
            issue_category = _category_from_issue(str(ruleid_val or "")) or None
        if issue_category:
            cells[ISSUE_CATEGORY_COLUMN] = issue_category
            non_empty = True

        root_cause = _cv("B", in_row)
        component = _cv("O", in_row)
        component = component.split("?")[0] if component else ""
        element = str(_cv("K", in_row))
        selector = _cv("I", in_row)
        issue_url_screenshot = _cv("R", in_row)
        help_url = _cv("D", in_row)

        howtofind = howtofix = ""
        if help_url is not None and str(help_url).strip() != "":
            # Compact here; placeholder 3 adds readable blank lines on load.
            howtofind = (f"HOW TO FIND - \nREPRODUCE:\n  Issue Type:\n"
                         f"{issue_category}\nComponent: {component}\n  URL: "
                         f"{issue_url_screenshot}\n  Element: {element}\n  "
                         f"Selector: {selector}\nWHERE TO FIX: Review the "
                         f"source/component shown above.\nVERIFY: Re-scan after "
                         f"the fix is deployed.\n\n")
            howtofix = (f"- HOW TO FIX - \nRule ID: {ruleid_val}\nImpact: "
                        f"{str(_cv('E', in_row))}\nIssue: {root_cause}\nHelp: "
                        f"{str(_cv('C', in_row))}\nRule URL: {help_url}\nFix:  "
                        f"apply the appropriate accessibility correction for "
                        f"{ruleid_val}")
        t_desc = (f"\n\nFix Owner:\nVendor\n\nWCAG Ref:\nRef:{first_id}\n\n"
                  f"WCAG:\n{first_id} {name}")
        cells["T"] = howtofind + howtofix + t_desc
        # Element column (H): the element with its selector appended at the end
        # in brackets, e.g. "button.cta [#main > button.cta]". (The Description
        # above keeps the separate Element: and Selector: lines.)
        elem_h = "" if str(element).strip().lower() in ("", "none") else str(element).strip()
        sel_h = "" if selector is None else str(selector).strip()
        if elem_h and sel_h:
            # The selector may already be wrapped in brackets (e.g.
            # ["#layout-drawer"]); only add a pair when it isn't, so we never
            # produce a double "[[...]]" (and CSS attribute selectors like
            # [role="button"] are kept as-is).
            sel_part = sel_h if (sel_h.startswith("[") and sel_h.endswith("]")) \
                else f"[{sel_h}]"
            cells["H"] = f"{elem_h} {sel_part}"
        else:
            cells["H"] = elem_h or sel_h

        # fixed summary columns
        cells["K"] = "Vendor"
        cells["L"] = "0"
        cells["M"] = "0"
        cells["P"] = "Pending"
        cells["Q"] = "PSGIN"
        cells["R"] = "Bug"
        cells["V"] = "kevin.murphy@ascendlearning.com"

        # standard tag (e.g. WCAG2A / WCAG21AA) -> F (overrides level)
        standard_id = []
        for token in str(_cv(WCAG_TAG_COLUMN, in_row)).split(","):
            if _WCAGAA_TAG_RE.match(token.strip().lower()):
                standard_id.append(token.strip().upper())
        if standard_id:
            cells["F"] = standard_id[0]

        # best-practice rows have no WCAG criterion; label the WCAG columns
        # "Best Practice" instead of leaving them blank.
        tags_lower = str(_cv(WCAG_TAG_COLUMN, in_row) or "").lower()
        if not cells.get(WCAG_ID_COLUMN) and "best-practice" in tags_lower:
            cells[WCAG_ID_COLUMN] = "Best Practice"      # N
            cells[WCAG_NAME_COLUMN] = "Best Practice"    # O
            if not cells.get("F"):
                cells["F"] = "Best Practice"             # WCAG Ver
            non_empty = True

        # Fallback: still no WCAG? Look up the rule id (Issue/G, from input A)
        # in Rule_ID_Mapping.xlsx and derive WCAG from the catalogue's tags.
        # Anything still unresolved is labelled "Best Practice" so the WCAG
        # columns are never left blank.
        if not cells.get(WCAG_ID_COLUMN):
            rid = str(_cv("A", in_row) or "").strip()
            cat_tags = _rule_catalogue().get(rid, "").replace(";", ",")
            cat_ids = _wcag_ids_from_cell(cat_tags) if cat_tags else []
            if cat_ids:
                fid = cat_ids[0]
                fid = cells[WCAG_ID_COLUMN] = (
                    fid.rsplit(".", 1)[0] if fid and fid[-1].isalpha() else fid)
                if fid in tags_map:
                    nm, lv = tags_map[fid]
                    cells[WCAG_NAME_COLUMN] = nm
                    cells[WCAG_LEVEL_COLUMN] = lv
                for tok in cat_tags.split(","):
                    if _WCAGAA_TAG_RE.match(tok.strip().lower()):
                        cells["F"] = tok.strip().upper()
                        break
            # final catch-all for custom/manual rules in no catalogue
            if not cells.get(WCAG_ID_COLUMN):
                cells[WCAG_ID_COLUMN] = "Best Practice"
                cells[WCAG_NAME_COLUMN] = "Best Practice"
            if not cells.get("F"):
                cells["F"] = "Best Practice"
            non_empty = True

        if parent_id:
            cells["W"] = _format_parent_id(parent_id)

        if not non_empty:
            continue

        cells["S"] = (f"Issue - {(out_row - 1):02d} - {str(cells.get('C', ''))} "
                      f"- {component} - {str(cells.get('G', ''))}")
        cells["A"] = f"Issue - {(out_row - 1):02d}"

        for out_col, value in cells.items():
            cell = out_ws[f"{out_col}{out_row}"]
            cell.value = normalize_xlsx_text(value)
            if out_col == "T":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        out_row += 1
        written += 1

    close_src()

    stem = _slug(out_name) if out_name else f"{_slug(path.stem)}_output"
    out = _outpath(stem)
    if not template_used:
        _format_standalone_sheet(out_ws, out_row - 1)
    out_wb.save(out)
    from services.excel_service import trim_trailing_blank_rows
    trim_trailing_blank_rows(out)
    from services.excel_service import highlight_summary_cells
    # highlighting removed per request: highlight_summary_cells(out)

    warnings = []
    if not template_used:
        warnings.append("No template at data/templates/Template_WCAG_Audit.xlsx "
                        "— generated a standalone workbook with the same columns.")
    if not tags_found:
        warnings.append(f"wcag_tags.txt not found — columns "
                        f"{WCAG_NAME_COLUMN}/{WCAG_LEVEL_COLUMN} left blank where "
                        f"no standard tag applied.")
    elif missing:
        shown = ", ".join(sorted(missing)[:12])
        more = "" if len(missing) <= 12 else f" (+{len(missing) - 12} more)"
        warnings.append(f"{len(missing)} id(s) not found in wcag_tags.txt: "
                        f"{shown}{more}")

    stats = {
        "written": written,
        "wcag_rows": wcag_rows,
        "matched": matched,
        "severity_rows": severity_rows,
        "sheet": sheet or "(auto)",
        "parent_id": parent_id or "",
        "template_used": template_used,
        "tags_found": tags_found,
        "target_sheet": TARGET_SHEET,
        "warnings": warnings,
    }
    # Pre-download verification: the four fields the user checks (and can fill
    # in place) before downloading — Rule ID (G/Issue), Description (T),
    # Selector (H/Element), Summary (S).
    stats["verify"] = _read_verify_preview(out, out_ws.title, written)
    return [("Audit workbook", out)], stats, _read_output_preview(out, out_ws.title)


def _format_standalone_sheet(ws, last_row: int) -> None:
    """Make the fallback workbook a clean, readable document."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hdr_fill = None  # no header highlighting
    hdr_font = Font(bold=True, size=11)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=False)
    thin = Side(style="thin", color="E4E8F0")
    border = Border(bottom=thin, right=thin)

    for col, width in _STANDALONE_WIDTHS.items():
        ws.column_dimensions[col].width = width
    for col in _STANDALONE_HEADERS:
        cell = ws[f"{col}1"]
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[1].height = 26

    for row in range(2, max(last_row, 1) + 1):
        for col in _STANDALONE_HEADERS:
            cell = ws[f"{col}{row}"]
            cell.border = border
            cell.alignment = (Alignment(vertical="top", wrap_text=True)
                              if col in _STANDALONE_WRAP else body_align)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:W{max(last_row, 1)}"


def _read_verify_preview(out_path: Path, sheet_name: str, n_rows: int) -> dict:
    """Read back the three fields the user verifies before downloading.

    Returns Rule ID (G/Issue), Issue Category (C) and Issue (D/Root Cause) for
    every written row, flagging any that is blank so the UI can highlight it and
    let the user fill it in place. Only the *n_rows* data rows that were
    actually written (rows 2 … n_rows+1) are read.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    FIELDS = [("G", "Rule ID"), ("C", "Issue Category"), ("D", "Issue")]
    last = 1 + max(0, int(n_rows))
    wb = load_workbook(out_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    sheet = ws.title
    idx = {label: column_index_from_string(col) - 1 for col, label in FIELDS}

    rows, empty_cells = [], 0
    if last >= 2:
        for rownum, r in enumerate(
                ws.iter_rows(min_row=2, max_row=last, values_only=True), start=2):
            values, empties = {}, []
            for _col, label in FIELDS:
                i = idx[label]
                v = r[i] if i < len(r) else None
                sv = "" if v is None else str(v).strip()
                if sv.lower() == "none":      # engine may write the literal "None"
                    sv = ""
                values[label] = sv
                if not sv:
                    empties.append(label)
            empty_cells += len(empties)
            rows.append({"row": rownum, "values": values, "empty": empties})
    wb.close()
    return {
        "columns": [label for _col, label in FIELDS],
        "field_cols": {label: col for col, label in FIELDS},
        "sheet": sheet,
        "rows": rows,
        "total": len(rows),
        "empty_cells": empty_cells,
        "empty_rows": sum(1 for r in rows if r["empty"]),
    }


def _read_output_preview(out_path: Path, sheet_name: str, limit: int = 200) -> dict:
    """Read back the populated sheet for the UI preview table."""
    from openpyxl import load_workbook
    wb = load_workbook(out_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or ()
    max_c = len(header)
    data_rows = []
    for r in rows_iter:
        data_rows.append(r)
        if len(data_rows) >= limit:
            break
    wb.close()
    # keep only columns that have a header or any data
    keep = []
    for i in range(max_c):
        label = header[i]
        has_data = any(i < len(r) and r[i] not in (None, "") for r in data_rows)
        if (label not in (None, "")) or has_data:
            keep.append(i)
    from openpyxl.utils import get_column_letter
    columns = [str(header[i]) if header[i] not in (None, "")
               else get_column_letter(i + 1) for i in keep]
    rows = []
    for r in data_rows:
        rows.append({columns[j]: ("" if (keep[j] >= len(r) or r[keep[j]] is None)
                                   else str(r[keep[j]])) for j in range(len(keep))})
    # total data rows
    total = max(0, (ws.max_row or 1) - 1) if hasattr(ws, "max_row") else len(rows)
    return {"columns": columns, "rows": rows, "total": len(rows), "shown": len(rows)}
