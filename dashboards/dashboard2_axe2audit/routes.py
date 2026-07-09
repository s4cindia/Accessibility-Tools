"""Placeholder 2 — Generate Axe → Audit Excel.

Operation page + conversion API, plus the shared template-management and
worksheet-picker endpoints used by the operation pages. The blueprint name
("axe_to_excel") is unchanged so existing url_for() calls keep working.
"""
from __future__ import annotations

from flask import Blueprint, request

from common.helpers import ok, fail, feature_response, render_operation
from common.file_manager import save_uploads
from common.logger import get_logger
from common.exceptions import ValidationError
from services import feature_service
from dashboards.dashboard2_axe2audit import converter, services, validation
from dashboards.dashboard2_axe2audit import constants as C

log = get_logger("sde.axe_to_excel")

axe_to_excel_bp = Blueprint(C.BLUEPRINT_NAME, __name__)


# ---- shared template management (used by placeholder 2 + 3) ----------------
@axe_to_excel_bp.route("/api/template/status")
def template_status():
    st = feature_service.template_status()
    return ok(found=st["found"], name=feature_service.WCAG_TEMPLATE.name)


@axe_to_excel_bp.route("/api/template/upload", methods=["POST"])
def template_upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return fail("No template file selected.")
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return fail("The template must be an Excel .xlsx (or .xlsm) file.")
    try:
        feature_service.TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        f.save(str(feature_service.WCAG_TEMPLATE))
        from openpyxl import load_workbook
        load_workbook(feature_service.WCAG_TEMPLATE).close()
    except Exception as exc:  # noqa: BLE001
        try:
            feature_service.WCAG_TEMPLATE.unlink()
        except OSError:
            pass
        log.exception("template upload failed")
        return fail(f"Could not use that template: {exc}")
    return ok(found=True, name=feature_service.WCAG_TEMPLATE.name, uploaded=f.filename)


@axe_to_excel_bp.route("/api/template/clear", methods=["POST"])
def template_clear():
    try:
        if feature_service.WCAG_TEMPLATE.exists():
            feature_service.WCAG_TEMPLATE.unlink()
    except OSError as exc:
        return fail(f"Could not remove template: {exc}")
    return ok(found=False, name=feature_service.WCAG_TEMPLATE.name)


# ---- the placeholder-2 operation -------------------------------------------
@axe_to_excel_bp.route(C.PAGE_URL)
def page():
    return render_operation(C.PAGE_SLUG)


@axe_to_excel_bp.route(C.API_CONVERT, methods=["POST"])
def axe2excel():
    f = request.files.get("file")
    sheet = (request.form.get("sheet") or "").strip() or None
    out_name = (request.form.get("out_name") or "").strip()
    try:
        parent_id = validation.require_parent_id(request.form.get("parent_id"))
        paths = save_uploads([f])
        outputs, stats, preview = converter.convert(
            paths[0], sheet=sheet, parent_id=parent_id, out_name=out_name)
    except ValidationError as exc:
        return fail(str(exc))
    except ValueError as exc:
        return fail(str(exc))
    except Exception as exc:  # noqa: BLE001
        log.exception("axe2excel failed")
        return fail(f"Conversion failed: {exc}")
    return feature_response(outputs, stats, preview)


@axe_to_excel_bp.route("/api/feature/axe2excel/verify-save", methods=["POST"])
def axe2excel_verify_save():
    """Write the user's inline edits to the three verify fields — Rule ID (G),
    Issue Category (C), Issue/Root Cause (D) — back into the generated audit
    workbook, so the downloaded file carries the filled-in values.

    Body: {file, sheet, edits:[{row, col, value}]}. Only the three editable
    columns are accepted, and only files inside this session's export folder.
    """
    from config import config
    from utils import file_utils

    body = request.get_json(silent=True) or {}
    name = (body.get("file") or "").strip()
    sheet = (body.get("sheet") or "").strip()
    edits = body.get("edits") or []
    if not name:
        return fail("Missing file name.")

    folder = file_utils.session_subdir(config.export_dir)
    target = (folder / name).resolve()
    # Stay inside the session's own export folder (no path traversal).
    if not str(target).startswith(str(folder.resolve())) or not target.exists():
        return fail("File not found.", 404)
    if not edits:
        return ok(saved=0, file=name)

    ALLOWED_COLS = {"G", "C", "D"}
    try:
        from openpyxl import load_workbook
        from services.excel_service import normalize_xlsx_text
        wb = load_workbook(target)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        saved = 0
        for e in edits:
            col = str(e.get("col") or "").strip().upper()
            try:
                row = int(e.get("row") or 0)
            except (TypeError, ValueError):
                continue
            if col not in ALLOWED_COLS or row < 2:
                continue
            val = e.get("value")
            ws[f"{col}{row}"].value = normalize_xlsx_text("" if val is None else str(val))
            saved += 1
        wb.save(target)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        log.exception("verify-save failed")
        return fail(f"Could not save your edits: {exc}")
    return ok(saved=saved, file=name)


@axe_to_excel_bp.route(C.API_SHEETS, methods=["POST"])
def feature_sheets():
    f = request.files.get("file")
    try:
        paths = save_uploads([f])
        sheets = services.sheets_for(paths[0])
    except ValueError as exc:
        return fail(str(exc))
    except Exception as exc:  # noqa: BLE001
        return fail(f"Could not read sheets: {exc}")
    return ok(sheets=sheets, file=paths[0].name)
