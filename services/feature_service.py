"""Business logic for the four dashboard feature operations.

Reworked to support the workflows shown in the legacy desktop tool while
keeping a clean web/service split:

  * sheet detection + selection            (axe2excel)
  * Parent ID applied to every output row  (axe2excel, downloadable)
  * "Media Inventory" sheet + type filter  (downloadable)
  * preview rows returned to the UI         (all)
  * WCAG / severity / priority breakdowns   (vpat)
  * multi-format export: xlsx, csv, VPAT    (vpat)
  * optional Template_WCAG_Audit.xlsx fill  (axe2excel, if present)

Each public function returns a 3-tuple ``(outputs, stats, preview)`` where
``outputs`` is a list of ``(label, Path)``. Validation problems raise
``ValueError`` with a user-facing message.
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from pathlib import Path

import polars as pl

from config import config
from services import export_service as _ex
from services.excel_service import normalize_xlsx_text, read_sheet

# Where an optional audit template can be dropped in to enable template output.
TEMPLATE_DIR = config.data_dir / "templates"
WCAG_TEMPLATE = TEMPLATE_DIR / "Template_WCAG_Audit.xlsx"
# Separate template used ONLY by Placeholder 3's "Export using delivery template".
WCAG_DELIVERY_TEMPLATE = TEMPLATE_DIR / "Template_WCAG_Delivery_1.3.xlsx"
# Boilerplate used ONLY by Placeholder 3's "Import Digital Asset": three audit
# rows (Word, PDF, PowerPoint) in the 23-column format. The static A–W values
# are read from this file — never hardcoded.
DIGITAL_ASSET_TEMPLATE = TEMPLATE_DIR / "Digital_Asset_Template.xlsx"
# Boilerplate used ONLY by Placeholder 3's "Import Alt Text": one audit row in
# the 23-column format. Static A–W values are read from this file.
ALT_TEXT_TEMPLATE = TEMPLATE_DIR / "Alt_Text_Template.xlsx"



def _read_table(path: str | Path, sheet: str | None = None) -> pl.DataFrame:
    """Read a worksheet (or csv/tsv) into an all-Utf8 Polars frame."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        from services.csv_service import decode_csv_bytes
        sep = "\t" if suffix == ".tsv" else ","
        # CSVs exported from Excel/Windows are frequently cp1252/latin-1 rather
        # than UTF-8, which makes a plain pl.read_csv raise "invalid utf-8
        # sequence". decode_csv_bytes returns clean UTF-8 so any source encoding
        # merges successfully.
        df = pl.read_csv(io.BytesIO(decode_csv_bytes(path)), separator=sep,
                         infer_schema_length=0, truncate_ragged_lines=True,
                         ignore_errors=True)
    else:
        df = read_sheet(path, sheet=sheet)
    df = df.rename({c: str(c).strip() for c in df.columns})
    if df.width:
        df = df.with_columns([pl.col(c).cast(pl.Utf8, strict=False) for c in df.columns])
    return df


def list_sheets_for(path: str | Path) -> list[str]:
    """Return worksheet names (xlsx/xls); empty list for csv/tsv."""
    path = Path(path)
    if path.suffix.lower() in (".csv", ".tsv"):
        return []
    try:
        from services.excel_service import list_sheets
        return list_sheets(path)
    except Exception:  # noqa: BLE001
        return []


def _find_col(df: pl.DataFrame, candidates: list[str]) -> str | None:
    lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    for cand in candidates:
        for lc, orig in lower.items():
            if cand.lower() in lc:
                return orig
    return None


def _col_or_blank(df: pl.DataFrame, col: str | None) -> pl.Expr:
    if col and col in df.columns:
        return pl.col(col).cast(pl.Utf8, strict=False).fill_null("")
    return pl.lit("")


def _outpath(stem: str, ext: str = ".xlsx") -> Path:
    return _ex._outpath(config.export_dir, stem, ext)


def _ts() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _slug(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_.\-]+", "_", (text or "").strip())
    return text.strip("_") or "output"


def _preview(df: pl.DataFrame, limit: int = 200) -> dict:
    head = df.head(limit)
    # Some source files carry the OOXML escape "_x000D_" as literal text (Excel
    # readers don't reverse it), which would otherwise show verbatim in the
    # preview. Normalise line endings for DISPLAY only — the underlying frame,
    # de-duplication, counts and column ordering are untouched.
    rows = [{k: normalize_xlsx_text(v) for k, v in row.items()}
            for row in head.to_dicts()]
    return {
        "columns": head.columns,
        "rows": rows,
        "total": df.height,
        "shown": head.height,
    }


def template_status() -> dict:
    return {"found": WCAG_TEMPLATE.exists(), "path": str(WCAG_TEMPLATE)}


_AUDIT_HEADERS_FALLBACK = [
    "ID", "Priority", "Issue Category", "Root Cause", "Component", "WCAG Ver",
    "Issue", "Element", "Issue URL", "Resolution", "Fix Owner", "Instances",
    "Pages", "WCAG Ref", "WCAG", "Verified", "Project Key", "Issue Type",
    "Summary", "Description", "Severity", "Assignee", "Parent",
]


def audit_headers() -> list[str]:
    """The 23 column headers a delivery import must match — read from the audit
    template if present, else the built-in list."""
    if WCAG_TEMPLATE.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(WCAG_TEMPLATE, read_only=True)
            ws = next((wb[n] for n in wb.sheetnames
                       if "all issues" in n.lower() or n.strip().startswith("2")),
                      wb.active)
            hdr = [str(ws.cell(1, c).value).strip()
                   for c in range(1, (ws.max_column or 0) + 1)
                   if ws.cell(1, c).value not in (None, "")]
            wb.close()
            if hdr:
                return hdr
        except Exception:  # noqa: BLE001
            pass
    return list(_AUDIT_HEADERS_FALLBACK)


def validate_audit_columns(columns) -> str | None:
    """Return a human-readable reason if `columns` don't match the audit format
    (the A–W headers), else None. Used to gate placeholder-3 imports.

    Column C must be exactly 'Issue Category' — files using any other label for
    that column (e.g. 'Issue Type') are rejected so the loaded/exported data is
    consistently 'Issue Category'."""
    from openpyxl.utils import get_column_letter

    def _accepts(expected_lower: str, got_lower: str) -> bool:
        return expected_lower == got_lower

    expected = audit_headers()
    got = [str(c).strip() for c in columns]
    exp_lower = [h.lower() for h in expected]
    got_lower = [c.lower() for c in got]
    if len(got) != len(expected):
        missing = [h for h in expected
                   if not any(_accepts(h.lower(), g) for g in got_lower)]
        extra = [c for c in got
                 if not any(_accepts(e, c.lower()) for e in exp_lower)]
        parts = [f"expected {len(expected)} columns (A–{get_column_letter(len(expected))}) "
                 f"but the file has {len(got)}"]
        if missing:
            parts.append("missing: " + ", ".join(missing))
        if extra:
            parts.append("unexpected: " + ", ".join(extra[:8]))
        return "; ".join(parts)
    mism = [(i + 1, expected[i], got[i]) for i in range(len(expected))
            if not _accepts(exp_lower[i], got_lower[i])]
    if mism:
        details = "; ".join(f"column {get_column_letter(i)} should be '{e}', found '{g}'"
                            for i, e, g in mism[:6])
        more = "" if len(mism) <= 6 else f" (+{len(mism) - 6} more)"
        return f"column names don't match the audit template — {details}{more}"
    return None


def normalize_issue_category_header(df: pl.DataFrame) -> pl.DataFrame:
    """Audit column C must be 'Issue Category', but some files label it
    'Issue Type' (which is also the legitimate name of column R). Rename ONLY the
    3rd column to 'Issue Category' when it reads 'Issue Type', by position, so
    column R is untouched. Returns the frame unchanged otherwise."""
    if df.width < 3 or str(df.columns[2]).strip().lower() != "issue type":
        return df
    names = list(df.columns)
    names[2] = "Issue Category"
    return df.select([pl.nth(i).alias(n) for i, n in enumerate(names)])


# Section labels that get a blank line before them in the structured
# (HOW TO FIND / HOW TO FIX) Description, for readability.
_DESC_BLANK_BEFORE = (
    "issue type:", "component:", "url:", "element:", "selector:", "- how to fix",
    "rule id:", "issue:", "help:", "rule url:", "fix:", "fix owner:",
    "wcag ref:", "wcag:",
)


def _space_one_description(text: str) -> str:
    """Insert a blank line before each labelled section of a structured
    ('HOW TO FIND … HOW TO FIX …') Description so it's readable in the grid.
    Idempotent (existing blank lines are first dropped, then re-inserted). Only
    touches the structured format (identified by 'REPRODUCE:'); other formats —
    the digital-asset labelled format, tail-only rows — are returned unchanged."""
    if not text or "REPRODUCE:" not in str(text).upper():
        return text
    lines = [ln for ln in str(text).split("\n") if ln.strip() != ""]
    out: list[str] = []
    for ln in lines:
        s = ln.strip().lower()
        if any(s.startswith(t) for t in _DESC_BLANK_BEFORE) and out and out[-1] != "":
            out.append("")
        out.append(ln)
    return "\n".join(out)


def space_description_sections(df: pl.DataFrame) -> pl.DataFrame:
    """Apply :func:`_space_one_description` to every 'Description' cell. Used by
    the Placeholder-3 (delivery) loader so a freshly-loaded sheet's Descriptions
    are readable. No-op when there's no Description column or no rows."""
    col = _exact_col(df, "Description")
    if not col or df.height == 0:
        return df
    spaced = (df.get_column(col).cast(pl.Utf8, strict=False)
                .map_elements(_space_one_description, return_dtype=pl.Utf8))
    return df.with_columns(spaced.alias(col))


def _exact_col(df: pl.DataFrame, name: str) -> str | None:
    """Return the column whose (stripped, case-insensitive) name equals `name`."""
    target = name.strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == target:
            return c
    return None


def _category_from_issue(value) -> str:
    """Turn an Issue value into an Issue Category: hyphens -> spaces, collapse
    whitespace, Title Case. e.g. 'html-has-lang' -> 'Html Has Lang'."""
    s = re.sub(r"-+", " ", str(value if value is not None else ""))
    s = re.sub(r"\s+", " ", s).strip()
    return s.title()


def fill_issue_category(df: pl.DataFrame) -> pl.DataFrame:
    """Fill any blank 'Issue Category' cell from that row's 'Issue' value, with
    hyphens turned into spaces and Title-Cased (e.g. 'html-has-lang' ->
    'Html Has Lang'). Cells that already have a value are left untouched.

    Used by the Placeholder-3 delivery editor both when the sheet is first
    loaded (so the grid shows the filled categories immediately) and again as a
    safety net at export time. Returns the frame unchanged if either column is
    missing or there are no rows."""
    issue_cat = _exact_col(df, "Issue Category")
    issue = _exact_col(df, "Issue")
    if not (issue_cat and issue) or df.height == 0:
        return df
    blank = (pl.col(issue_cat).is_null()
             | (pl.col(issue_cat).cast(pl.Utf8, strict=False)
                  .str.strip_chars().fill_null("") == ""))
    cat_from_issue = (df.get_column(issue)
                        .cast(pl.Utf8, strict=False).fill_null("")
                        .map_elements(_category_from_issue, return_dtype=pl.Utf8))
    return (df.with_columns(cat_from_issue.alias("__cat_from_issue"))
              .with_columns(pl.when(blank).then(pl.col("__cat_from_issue"))
                              .otherwise(pl.col(issue_cat).cast(pl.Utf8, strict=False))
                              .alias(issue_cat))
              .drop("__cat_from_issue"))


def strip_description_hyphen(df: pl.DataFrame) -> pl.DataFrame:
    """Remove a leading hyphen (and any surrounding spaces) from every
    'Description' cell, e.g. '- HOW TO FIND -…' -> 'HOW TO FIND -…'. A leading
    hyphen makes Excel treat the text as a formula, so it can't be edited.

    Applied by the Placeholder-3 delivery editor when the sheet loads (so the
    grid never shows a leading hyphen) and again at export. Returns the frame
    unchanged if there is no 'Description' column."""
    desc = _exact_col(df, "Description")
    if not desc or df.height == 0:
        return df
    return df.with_columns(
        pl.col(desc).cast(pl.Utf8, strict=False)
          .str.replace(r"^\s*-+\s*", "").alias(desc)
    )


EXCEL_ERROR_VALUES = {"#DIV/0!", "#VALUE!", "#REF!", "#N/A", "#NUM!", "#NAME?",
                      "#NULL!", "#SPILL!", "#CALC!"}


def _excel_error_cells(df):
    """Return {column: [1-based row numbers]} for any cell holding an Excel
    error literal (e.g. #REF!, #DIV/0!)."""
    import polars as _pl
    hits = {}
    upper = {e.upper() for e in EXCEL_ERROR_VALUES}
    for c in df.columns:
        vals = (df.select(_pl.col(c).cast(_pl.Utf8, strict=False)
                            .str.strip_chars().str.to_uppercase().alias("v"))
                  .get_column("v").to_list())
        rows = [i + 1 for i, v in enumerate(vals) if v in upper]
        if rows:
            hits[c] = rows
    return hits


WCAG_TAGS = TEMPLATE_DIR / "wcag_tags.txt"


def load_wcag_tags(path: Path) -> dict:
    mapping = {}
    if not path or not Path(path).exists():
        return mapping
    with open(path, encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3 and parts[0]:
                mapping[parts[0]] = (parts[1], parts[2])
    return mapping
